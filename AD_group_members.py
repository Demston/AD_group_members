"""Выборка участников группы домена из Active Directory"""

from ldap3 import Server, Connection, SUBTREE
import openpyxl
from datetime import datetime

# Определяем сервер и считываем учётные данные
AD_SERVER = 'dc1.domain.com'
AD_SERVER_RESERV = 'dc2.domain.com'
domain_name = 'DOMAIN'
# файл ini  логин:пароль
your_name = open('ini', encoding='UTF-8').read().split(':')[0]
AD_USER = f'{domain_name}\{your_name}'
AD_PASSWORD = open('ini', encoding='UTF-8').read().split(':')[1]
AD_SEARCH_TREE = 'dc=roscap,dc=com'

# Проверяем и устанавливаем соединение с сервером
server = ''
if Connection(AD_SERVER).bind() is True:
    server = Server(AD_SERVER)
else:
    server = Server(AD_SERVER_RESERV)
conn = Connection(server, user=AD_USER, password=AD_PASSWORD)
conn.bind()

print(f'Выборка участников группы домена {domain_name}')
group_name = input(f'Введи название группы, например: example_group\n\n')


def create_table(nms, lgns):
    """Функция записи результатов в таблицу"""
    current_datetime = datetime.now()
    datetime_string = current_datetime.strftime('%d%m%Y_%H%M%S')
    wb = openpyxl.Workbook()
    wb['Sheet'].title = group_name
    ws = wb.active
    ws['B1'] = 'Пользователи, состоящие в группе Active Directory'
    ws['B2'] = f'{domain_name}|'+group_name+'   на   '+current_datetime.strftime('%d.%m.%Y')
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 60
    count1 = 3
    for lgn in lgns:
        count1 += 1
        ws[f'A{count1}'] = lgn
    count2 = 3
    for nm in nms:
        count2 += 1
        ws[f'B{count2}'] = nm
    wb.save(group_name+' '+datetime_string+'.xlsx')


while True:
    try:    # Поиск группы с атрибутами: фильтр (тип, имя группы), дерево, атрибут - именно в нём перечислены юзеры)
        group_filter = f"(&(objectClass=group)(&(cn={group_name})))"
        conn.search(AD_SEARCH_TREE, group_filter, SUBTREE, attributes=['member'])
        # Деление по спискам (общий, имена, логины)
        members_list = []
        names = []
        logins = []
        for member_attribute in conn.entries:
            for member in member_attribute:
                for m in member:
                    name = m[3:].split(',')[0]
                    members_list.append(name)
        # Поиск по логинам, взятым из одного из атрибутов (логин, с деревом, возврат атрибутов)
        members_list.sort()
        for user in members_list:
            conn.search(AD_SEARCH_TREE, f'(&(objectCategory=Person)(|(cn={user})))', SUBTREE,
                        attributes=['cn', 'sAMAccountName', 'UserAccountControl'])
            # Добавление в списки, вывод на экран
            for entry in conn.entries:
                if entry.UserAccountControl == 512:     # Проверка статуса УЗ: Вкл.
                    names.append(str(entry.cn))
                    logins.append(str(entry.sAMAccountName))
                    print(entry.cn, '|', entry.sAMAccountName)

        create_table(names, logins)
        print('\nГотово. Файл сохранён в таблицу (в папке со скриптом)\n')

        if input('Нажми что-нибудь, чтобы выйти\n') == '':
            break
        else:
            break

    except:
        print('\nЧто-то пошло не так\n')
